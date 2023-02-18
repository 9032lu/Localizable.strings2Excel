# -*- coding:utf-8 -*-

import os
from optparse import OptionParser
from StringsFileUtil import StringsFileUtil
import openpyxl
import time
from Log import Log
import shutil
# Add command option


def addParser():
    parser = OptionParser()

    parser.add_option("-f", "--stringsDir",
                      help=".strings files directory.",
                      metavar="stringsDir")

    parser.add_option("-t", "--targetDir",
                      help="The directory where the excel(.xls) files will be saved.",
                      metavar="targetDir")

    parser.add_option("-e", "--excelStorageForm",
                      type="string",
                      default="single",
                      help="The excel(.xls) file storage forms including single(single file), multiple(multiple files), default is multiple.",
                      metavar="excelStorageForm")

    (options, _) = parser.parse_args()

    return options

#  convert .strings files to single xls file


def convertToSingleFile(stringsDir, targetDir):
    destDir = targetDir
    # destDir = targetDir + "/strings-files-to-xls_" + \
    #     time.strftime("%Y%m%d_%H%M%S")
    if not os.path.exists(destDir):
        os.makedirs(destDir)

    # Create xls sheet
    i = 1
    for _, dirnames, _ in os.walk(stringsDir):
        if len(dirnames)==0 and i==1:
            Log.info('-----------没有可用的.lproj文件')

        if len(dirnames) == 0 : continue
        Log.info(dirnames)
        
        lprojDirs = [di for di in dirnames if di.endswith(".lproj")]
        if len(dirnames)==0 and i==1:
            Log.info('-----------没有可用的.lproj文件')
        i+=1
        for dirname in lprojDirs:
            for _, _, filenames in os.walk(stringsDir+'/'+dirname):
                stringsFiles = [
                    fi for fi in filenames if fi.endswith(".strings")]
                for stringfile in stringsFiles:
                    fileName = stringfile.replace(".strings", "")
                    filePath = destDir + "/" + fileName + ".xlsx"
                    if not os.path.exists(filePath):
                        workbook = openpyxl.Workbook()
                        ws = workbook.create_sheet(fileName,0)
                        index = 0
                        for dirname in dirnames:
                            if index == 0:
                                ws.cell(row=1,column = 1).value='keyName'
                            countryCode = dirname.replace(".lproj", "")
                            ws.cell(row=1,column = index+2).value=countryCode

                            path = stringsDir+'/' + dirname + '/' + stringfile
                            (keys, values) = StringsFileUtil.getKeysAndValues(
                                path)
                            for x in range(len(keys)):
                                key = keys[x]
                                value = values[x]
                                if (index == 0):
                                    ws.cell(row=x+2,column=1).value=key
                                    ws.cell(row=x+2,column=2).value=value
                                else:
                                    ws.cell(row=x+2,column=index+2).value=value
                            index += 1
                        workbook.save(filePath)
    
    print ("Convert %s successfully! you can see xls file in %s" % (
        stringsDir, destDir))
    Log.info('转换完成，速度杠杠的！！！')


#  convert .strings files to multiple xls files


def convertToMultipleFiles(stringsDir, targetDir):
    destDir = targetDir + "/strings-files-to-xls_" + \
        time.strftime("%Y%m%d_%H%M%S")
    if not os.path.exists(destDir):
        os.makedirs(destDir)

    for _, dirnames, _ in os.walk(stringsDir):
        lprojDirs = [di for di in dirnames if di.endswith(".lproj")]
        for dirname in lprojDirs:
            workbook = openpyxl.Workbook()
            print(dirname)
            for _, _, filenames in os.walk(stringsDir+'/'+dirname):
                stringsFiles = [
                    fi for fi in filenames if fi.endswith(".strings")]
                for stringfile in stringsFiles:
                    ws = workbook.create_sheet(stringfile)
                    path = stringsDir+dirname+'/' + stringfile
                    (keys, values) = StringsFileUtil.getKeysAndValues(
                        path)
                    for keyIndex in range(len(keys)):
                        key = keys[keyIndex]
                        value = values[keyIndex]
                        inde = keyIndex + 1
                        ws.cell(row= inde,column=1).value = key
                        ws.cell(row= inde,column=2).value = value

            filePath = destDir + "/" + dirname.replace(".lproj", "") + ".xlsx"
            workbook.save(filePath)

    print ("Convert %s successfully! you can see xls file in %s" % (
        stringsDir, destDir))

# Start convert .strings files to xls


def startConvert(options):
    stringsDir = options.stringsDir
    targetDir = options.targetDir

    print ("Start converting")

    if stringsDir is None:
        print (".strings files directory can not be empty! try -h for help.")
        return

    if targetDir is None:
        print ("Target file directory can not be empty! try -h for help.")
        return

    if options.excelStorageForm == "single":
        convertToSingleFile(stringsDir, targetDir)
    else:
        convertToMultipleFiles(stringsDir, targetDir)

# 脚本执行放开main
# def main():
#     options = addParser()
#     startConvert(options)

# main()

class Strings2Xlsx:
    '工具用来执行方法'
    @staticmethod
    def startConvertStringToXlsx(fileDir,targetDir):
        Log.info(f'xlsx路径:{fileDir}')
        Log.info(f'保存路径:{targetDir}')
        if fileDir == targetDir:
            Log.info('路径一致会被清空！！！！')
            return
        if not os.path.exists(targetDir):
            os.makedirs(targetDir)
        else:
            shutil.rmtree(targetDir)
            os.makedirs(targetDir)

        convertToSingleFile(fileDir,targetDir)