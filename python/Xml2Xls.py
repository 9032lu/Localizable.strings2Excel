# -*- coding:utf-8 -*-

import os
from optparse import OptionParser
from XmlFileUtil import XmlFileUtil
import openpyxl
from Log import Log
import time
import shutil

def addParser():
    parser = OptionParser()

    parser.add_option("-f", "--fileDir",
                      help="strings.xml files directory.",
                      metavar="fileDir")

    parser.add_option("-t", "--targetDir",
                      help="The directory where the xls files will be saved.",
                      metavar="targetDir")

    parser.add_option("-e", "--excelStorageForm",
                      type="string",
                      default="single",
                      help="The excel(.xls) file storage forms including single(single file), multiple(multiple files), default is multiple.",
                      metavar="excelStorageForm")

    (options, args) = parser.parse_args()
    Log.info("options: %s, args: %s" % (options, args))

    return options


def convertToMultipleFiles(fileDir, targetDir):
    destDir = genDestDir(targetDir)

    for _, dirnames, _ in os.walk(fileDir):
        valuesDirs = [di for di in dirnames if di.startswith("values")]
        for dirname in valuesDirs:
            workbook = pyExcelerator.Workbook()
            for _, _, filenames in os.walk(fileDir+'/'+dirname):
                xmlFiles = [fi for fi in filenames if fi.endswith(".xml")]
                for xmlfile in xmlFiles:
                    ws = workbook.add_sheet(xmlfile)
                    path = fileDir+'/'+dirname+'/' + xmlfile
                    (keys, values) = XmlFileUtil.getKeysAndValues(path)
                    for keyIndex in range(len(keys)):
                        key = keys[keyIndex]
                        value = values[keyIndex]
                        ws.write(keyIndex, 0, key)
                        ws.write(keyIndex, 1, value)
            filePath = destDir + "/" + getCountryCode(dirname) + ".xls"
            workbook.save(filePath)
    print ("Convert %s successfully! you can see xls file in %s" % (
        fileDir, destDir))


def convertToSingleFile(fileDir, targetDir):
    destDir = genDestDir(targetDir)
    i = 0
    for _, dirnames, _ in os.walk(fileDir):
        if len(dirnames)==0 and i==0:
            Log.info('--没有可用xml')

        valuesDirs = [di for di in dirnames if di.startswith("values")]
        if i==0:Log.info(f'--valuesDirs:{valuesDirs}')

        if len(valuesDirs)==0 and i ==0:
            Log.info('---没有可用xml')
        i+=1
        for dirname in valuesDirs:
            print(f'dirnameL:{dirname}')
            for _, _, filenames in os.walk(fileDir+'/'+dirname):
                xmlFiles = [fi for fi in filenames if fi.endswith(".xml")]
                for xmlfile in xmlFiles:
                    fileName = xmlfile.replace(".xml", "")
                    filePath = destDir + "/" + fileName + ".xlsx"
                    if not os.path.exists(filePath):
                        workbook = openpyxl.Workbook()
                        ws = workbook.create_sheet(fileName,0)
                        index = 0
                        for dirname in dirnames:
                            print(f"---{dirname}")
                            if index == 0:
                                # ws.write(0, 0, 'keyName')
                                ws.cell(row=1,column=1).value = 'keyName'
                            countryCode = getCountryCode(dirname)
                            # ws.write(0, index+1, countryCode)
                            print(countryCode)
                            ws.cell(row=1,column=index+2).value = countryCode

                            path = fileDir+'/'+dirname+'/' + xmlfile
                            (keys, values) = XmlFileUtil.getKeysAndValues(path)
                            for x in range(len(keys)):
                                key = keys[x]
                                value = values[x]
                                # if (index == 0):
                                #     ws.write(x+1, 0, key)
                                #     ws.write(x+1, 1, value)
                                # else:
                                #     ws.write(x+1, index + 1, value)
                                if (index == 0):
                                    ws.cell(row=x+2,column=1).value=key
                                    ws.cell(row=x+2,column=2).value=value
                                else:
                                    ws.cell(row=x+2,column=index+2).value=value
                            index += 1
                        workbook.save(filePath)
    print ("Convert %s successfully! you can see xls file in %s" % (
        fileDir, destDir))
    Log.info('转换完成，速度杠杠的！！！')

def genDestDir(targetDir):
    destDir = targetDir
    # destDir = targetDir + "/xml-files-to-xls_" + \
        # time.strftime("%Y%m%d_%H%M%S")
    if not os.path.exists(destDir):
        os.makedirs(destDir)

    return destDir


def getCountryCode(dirname):
    code = 'en'
    dirSplit = dirname.split('values-')
    if len(dirSplit) > 1:
        code = dirSplit[1]
    return code


def startConvert(options):
    fileDir = options.fileDir
    targetDir = options.targetDir

    print ("Start converting")

    if fileDir is None:
        print ("strings.xml files directory can not be empty! try -h for help.")
        return

    if targetDir is None:
        print ("Target file path can not be empty! try -h for help.")
        return

    if options.excelStorageForm == "single":
        convertToSingleFile(fileDir, targetDir)
    else:
        convertToMultipleFiles(fileDir, targetDir)


# def main():
#     options = addParser()
#     startConvert(options)


# main()

class Xml2Xlsx:
    '工具用来执行方法'
    @staticmethod
    def convertToSingleFile(fileDir,targetDir):
        Log.info(f'xlsx路径:{fileDir}')
        Log.info(f'保存路径:{targetDir}')
        if fileDir == targetDir:
            Log.info('路径一致会被清空！！！！')
            return

        if not os.path.exists(targetDir):
            os.makedirs(targetDir)
        else:
            shutil.rmtree(targetDir)
            os.makedirs(targetDir)
        
        convertToSingleFile(fileDir,targetDir)